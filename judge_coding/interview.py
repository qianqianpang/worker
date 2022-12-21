# _*_coding : utf-8 _*_
# @Time : 2022/12/3 11:34
# @Author : 秦汉强
# @File : meeting_summary.py
# @Project : pythonProject4

import openpyxl
import re
import json

def run(filepath):
    workbook = openpyxl.load_workbook(filepath)
    sheet_list = workbook.sheetnames
    a = {}
    for sheet in sheet_list:
        sheet = workbook[sheet]
        a['主题'] = sheet.cell(row=2, column=3).value
        a['日期'] = re.findall('(.*?)日',sheet.cell(row=3, column=3).value,re.S)[0] + '日'
        a['开始时间'] = re.findall('日(.*?)~',sheet.cell(row=3, column=3).value,re.S)[0].strip()
        a['结束时间'] = re.findall('~(.*)',sheet.cell(row=3, column=3).value,re.S)[0].strip()
        a['地点'] = sheet.cell(row=3, column=10).value
        a['主持人'] = sheet.cell(row=4, column=3).value
        a['记录人'] = sheet.cell(row=4, column=10).value
        a['应出席人数'] = sheet.cell(row=20, column=3).value
        a['实际出席人数'] = sheet.cell(row=20, column=5).value
        a['缺席人数'] = sheet.cell(row=20, column=10).value
        a['出席率'] = str(round(int(a['实际出席人数']) / int(a['应出席人数']),3) * 100) + '%'
        list_a = []
        for row in range(7,20):
            a_1 = {}
            a_1['姓名'] = sheet.cell(row=row, column=3).value
            a_1['部门'] = sheet.cell(row=row,column=2).value
            if sheet.cell(row=row, column=4).value is None:
                a_1['签到'] = ''
            else:
                a_1['签到'] = sheet.cell(row=row, column=4).value
            if  sheet.cell(row=row, column=5).value is not None:
                a_1["出席状态"] = '缺席'
            else:
                a_1["出席状态"] = '出席'
            if  sheet.cell(row=row, column=6).value is not None:
                a_1["是否迟到"] = True
            else:
                a_1["是否迟到"] = False
            if  sheet.cell(row=row, column=7).value is not None:
                a_1["是否早退"] = True
            else:
                a_1["是否早退"] = False
            list_a.append(a_1)

        for row in range(7,11):
            a_2 = {}
            a_2['姓名'] = sheet.cell(row=row, column=10).value
            a_2['部门'] = sheet.cell(row=row,column=9).value
            if sheet.cell(row=row, column=11).value is None:
                c = 2
                a_2['签到'] = ''
            else:
                a_2['签到'] = sheet.cell(row=row, column=11).value
            if  sheet.cell(row=row, column=12).value is not None:
                a_2["出席状态"] = '缺席'
            else:
                a_2["出席状态"] = '出席'
            if  sheet.cell(row=row, column=13).value is not None:
                a_2["是否迟到"] = True
            else:
                a_2["是否迟到"] = False
            if  sheet.cell(row=row, column=14).value is not None:
                a_2["是否早退"] = True
            else:
                a_2["是否早退"] = False
            list_a.append(a_2)
        a["参会人详细信息"] = list_a

        a["各组出勤情况"] = {}
        xiaoshou1_num = 0
        xiaoshou1_chuqing = 0
        xiaoshou1_queqing = 0
        xiaoshou1_chidao = 0
        xiaoshou1_zaotui = 0
        xiaoshou2_num = 0
        xiaoshou2_chuqing = 0
        xiaoshou2_queqing = 0
        xiaoshou2_chidao = 0
        xiaoshou2_zaotui = 0
        xiaoshou3_num = 0
        xiaoshou3_chuqing = 0
        xiaoshou3_queqing = 0
        xiaoshou3_chidao = 0
        xiaoshou3_zaotui = 0
        yanfa1_num = 0
        yanfa1_chuqing = 0
        yanfa1_queqing = 0
        yanfa1_chidao = 0
        yanfa1_zaotui = 0
        yanfa2_num = 0
        yanfa2_chuqing = 0
        yanfa2_queqing = 0
        yanfa2_chidao = 0
        yanfa2_zaotui = 0
        chanpin_num = 0
        chanpin_chuqing = 0
        chanpin_queqing = 0
        chanpin_chidao = 0
        chanpin_zaotui = 0
        for yuangon_comtent in list_a:
            bumen_xinxi = yuangon_comtent['部门']
            if '销售1部' in bumen_xinxi:
                a["各组出勤情况"]['销售1部'] = {}
                xiaoshou1_num += 1    #应出勤人数
                if yuangon_comtent['出席状态'] == '出席':
                    xiaoshou1_chuqing += 1
                else:
                    xiaoshou1_queqing += 1
                if yuangon_comtent['是否迟到'] is True:
                    xiaoshou1_chidao += 1
                if yuangon_comtent['是否早退'] is True:
                    xiaoshou1_zaotui += 1
                a["各组出勤情况"]['销售1部']["应出勤人数"] = xiaoshou1_num
                a["各组出勤情况"]['销售1部']["实际出勤人数"] = xiaoshou1_chuqing
                a["各组出勤情况"]['销售1部']["出勤率"] = str(int(round(xiaoshou1_chuqing/xiaoshou1_num,3) * 100)) + '%'
                a["各组出勤情况"]['销售1部']["迟到比例"] = str(int(round(xiaoshou1_chidao/xiaoshou1_num,3) * 100)) + '%'
                a["各组出勤情况"]['销售1部']["早退比例"] = str(int(round(xiaoshou1_zaotui/xiaoshou1_num,3) * 100)) + '%'
            if '销售2部' in bumen_xinxi:
                a["各组出勤情况"]['销售2部'] = {}
                xiaoshou2_num += 1    #应出勤人数
                if yuangon_comtent['出席状态'] == '出席':
                    xiaoshou2_chuqing += 1
                else:
                    xiaoshou2_queqing += 1
                if yuangon_comtent['是否迟到'] is True:
                    xiaoshou2_chidao += 1
                if yuangon_comtent['是否早退'] is True:
                    xiaoshou2_zaotui += 1
                a["各组出勤情况"]['销售2部']["应出勤人数"] = xiaoshou2_num
                a["各组出勤情况"]['销售2部']["实际出勤人数"] = xiaoshou2_chuqing
                a["各组出勤情况"]['销售2部']["出勤率"] = str(int(round(xiaoshou2_chuqing/xiaoshou2_num,3) * 100)) + '%'
                a["各组出勤情况"]['销售2部']["迟到比例"] = str(int(round(xiaoshou2_chidao/xiaoshou2_num,3) * 100)) + '%'
                a["各组出勤情况"]['销售2部']["早退比例"] = str(int(round(xiaoshou2_zaotui/xiaoshou2_num,3) * 100)) + '%'
            if '销售3部' in bumen_xinxi:
                a["各组出勤情况"]['销售3部'] = {}
                xiaoshou3_num += 1    #应出勤人数
                if yuangon_comtent['出席状态'] == '出席':
                    xiaoshou3_chuqing += 1
                else:
                    xiaoshou3_queqing += 1
                if yuangon_comtent['是否迟到'] is True:
                    xiaoshou3_chidao += 1
                if yuangon_comtent['是否早退'] is True:
                    xiaoshou3_zaotui += 1
                a["各组出勤情况"]['销售3部']["应出勤人数"] = xiaoshou3_num
                a["各组出勤情况"]['销售3部']["实际出勤人数"] = xiaoshou3_chuqing
                a["各组出勤情况"]['销售3部']["出勤率"] = str(int(round(xiaoshou3_chuqing/xiaoshou3_num,3) * 100)) + '%'
                a["各组出勤情况"]['销售3部']["迟到比例"] = str(int(round(xiaoshou3_chidao/xiaoshou3_num,3) * 100)) + '%'
                a["各组出勤情况"]['销售3部']["早退比例"] = str(int(round(xiaoshou3_zaotui/xiaoshou3_num,3) * 100)) + '%'

            if '研发1部' in bumen_xinxi:
                a["各组出勤情况"]['研发1部'] = {}
                yanfa1_num += 1    #应出勤人数
                if yuangon_comtent['出席状态'] == '出席':
                    yanfa1_chuqing += 1
                else:
                    yanfa1_queqing += 1
                if yuangon_comtent['是否迟到'] is True:
                    yanfa1_chidao += 1
                if yuangon_comtent['是否早退'] is True:
                    yanfa1_zaotui += 1
                a["各组出勤情况"]['研发1部']["应出勤人数"] = yanfa1_num
                a["各组出勤情况"]['研发1部']["实际出勤人数"] = yanfa1_chuqing
                a["各组出勤情况"]['研发1部']["出勤率"] = str(int(round(yanfa1_chuqing/yanfa1_num,3) * 100)) + '%'
                a["各组出勤情况"]['研发1部']["迟到比例"] = str(int(round(yanfa1_chidao/yanfa1_num,3) * 100)) + '%'
                a["各组出勤情况"]['研发1部']["早退比例"] = str(int(round(yanfa1_zaotui/yanfa1_num,3) * 100)) + '%'

            if '研发2部' in bumen_xinxi:
                a["各组出勤情况"]['研发2部'] = {}
                yanfa2_num += 1    #应出勤人数
                if yuangon_comtent['出席状态'] == '出席':
                    yanfa2_chuqing += 1
                else:
                    yanfa2_queqing += 1
                if yuangon_comtent['是否迟到'] is True:
                    yanfa2_chidao += 1
                if yuangon_comtent['是否早退'] is True:
                    yanfa2_zaotui += 1
                a["各组出勤情况"]['研发2部']["应出勤人数"] = yanfa2_num
                a["各组出勤情况"]['研发2部']["实际出勤人数"] = yanfa2_chuqing
                a["各组出勤情况"]['研发2部']["出勤率"] = str(int(round(yanfa2_chuqing/yanfa2_num,3) * 100)) + '%'
                a["各组出勤情况"]['研发2部']["迟到比例"] = str(int(round(yanfa2_chidao/yanfa2_num,3) * 100)) + '%'
                a["各组出勤情况"]['研发2部']["早退比例"] = str(int(round(yanfa2_zaotui/yanfa2_num,3) * 100)) + '%'

            if '产品部' in bumen_xinxi:
                a["各组出勤情况"]['产品部'] = {}
                chanpin_num += 1    #应出勤人数
                if yuangon_comtent['出席状态'] == '出席':
                    chanpin_chuqing += 1
                else:
                    chanpin_queqing += 1
                if yuangon_comtent['是否迟到'] is True:
                    chanpin_chidao += 1
                if yuangon_comtent['是否早退'] is True:
                    chanpin_zaotui += 1
                a["各组出勤情况"]['产品部']["应出勤人数"] = chanpin_num
                a["各组出勤情况"]['产品部']["实际出勤人数"] = chanpin_chuqing
                a["各组出勤情况"]['产品部']["出勤率"] = str(int(round(chanpin_chuqing/chanpin_num,3) * 100)) + '%'
                a["各组出勤情况"]['产品部']["迟到比例"] = str(int(round(chanpin_chidao/chanpin_num,3) * 100)) + '%'
                a["各组出勤情况"]['产品部']["早退比例"] = str(int(round(chanpin_zaotui/chanpin_num,3) * 100)) + '%'
    a_json = json.dumps(a,ensure_ascii=False)
    with open('data.json',mode='w',encoding='utf-8') as f:
        f.write(a_json)
    #缺席
    quexi_content = ''
    quexi_account = 0
    cdzt_content = ''
    weiqiandao_content = ''

    for n in a['参会人详细信息']:
        if '出席' not in n['出席状态']:
            quexi_account += 1
            quexi_content += n['姓名'] + '(' + n['部门']+ ')' + '、'
        if n['是否早退'] is True or n['是否迟到'] is True :
            cdzt_content += n['姓名'] + '(' + n['部门'] + ')' + '\n'
        if len(n['签到']) == 0:
            weiqiandao_content += n['姓名'] + '、'
    content = """
会议{}于{} {}至{}进行。我整理了会议签到表，以下是会议参与情况请注意查看。
会议应出席{}人，实际出席{}人，其中{}（{}位）缺席会议。

由于部分同事来迟或提前离开，请会后将会议全程录屏抄送给他们：
{}

签到表上部分同事未来得及签名，可否通知他们来我这里补签，非常感谢。
未签到同事：{}
""".format(a['主题'],a['日期'],a['开始时间'],a['结束时间'],a['应出席人数'],a['实际出席人数'],quexi_content[:-1],quexi_account,cdzt_content[:-1],weiqiandao_content[:-1]).strip()
    with open('summary.txt',mode = 'w',encoding='utf-8') as f:
        f.write(content)
    # 请在此处实现代码，结果输出至data.json和summary.txt中

# 代码执行方式, 命令行模式下，进入代码所在根目录，然后执行: python meeting_summary.py
# 请确保下载的测试文件放在了和代码相同的目录下
if __name__ == '__main__':
    run("公司会议签到表.xlsx")