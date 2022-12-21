from openpyxl import load_workbook
import json

str_date = ""


# 处理参会人员信息
def handle_info(people_info):
    if people_info[2] is None:
        people_info[2] = ""
    if people_info[3] is None:
        people_info[3] = "出席"
    else:
        people_info[3] = "缺席"
    if people_info[4] is None:
        people_info[4] = False
    else:
        people_info[4] = True
    if people_info[5] is None:
        people_info[5] = False
    else:
        people_info[5] = True
    info = {"部门": people_info[0], "姓名": people_info[1], "签到": people_info[2], "出席状态": people_info[3],
            "是否迟到": people_info[4], "是否早退": people_info[5]}
    return info


# 写入summary
def write_file(result, path):
    absence_info = ""
    buqian_info = "\n签到表上部分同事未来得及签名，可否通知他们来我这⾥补签，⾮常感谢。\n未签到同事："
    late_earlyGo_info = ""
    for people in result["参会人详细信息"]:
        if people["出席状态"] == "缺席":
            absence_info += people["姓名"]
            absence_info += "("
            absence_info += people["部门"]
            absence_info += ")、"
    absence_info = absence_info[:-1]
    absence_info += "("
    absence_info += str(result["缺席人数"])
    absence_info += "位)缺席会议。\n\n"

    for people in result["参会人详细信息"]:
        if people["是否早退"] == True or people["是否迟到"] == True:
            late_earlyGo_info += people["姓名"]
            late_earlyGo_info += "("
            late_earlyGo_info += people["部门"]
            late_earlyGo_info += ")\n"

    for people in result["参会人详细信息"]:
        if people["签到"] == "" and people["出席状态"] == "出席":
            buqian_info += people["姓名"]
            buqian_info += "、"
    if buqian_info != "\n签到表上部分同事未来得及签名，可否通知他们来我这⾥补签，⾮常感谢。\n未签到同事：":
        buqian_info = buqian_info[:-1]

    with open(path, 'w', encoding='utf-8') as f:
        global str_date
        content = "会议" \
                  + result["主题"] \
                  + "于" \
                  + str_date \
                  + result["开始时间"] \
                  + "至" \
                  + result["结束时间"] \
                  + "进行。我整理了会议签到表，以下是会议参与情况请注意查看。\n\n"
        content += "会议应出席" \
                   + str(result["应出席人数"]) \
                   + "⼈，实际出席" \
                   + str(result["实际出席人数"]) \
                   + "⼈，其中" \
                   + absence_info
        content += "由于部分同事来迟或提前离开，请会后将会议全程录屏抄送给他们：\n"
        content += late_earlyGo_info
        content += buqian_info
        f.write(content)


def format_date(date_str):
    date_str = date_str.replace("年", "-")
    date_str = date_str.replace("月", "-")
    date_str = date_str.replace("日", "")
    res_list = date_str.split('-')
    date_str = '' + res_list[0] + '-'
    if int(res_list[1]) < 10:
        date_str += '0'
        date_str += res_list[1]
        date_str += '-'
    date_str += res_list[2]
    return date_str


def run(filepath):
    actual_amount = 0
    absence_amount = 0
    total_amount = 0
    result = {}
    workbook = load_workbook(filename=filepath)
    sheet = workbook["Sheet1"]
    result["主题"] = sheet["C2"].value
    global str_date
    str_date = sheet["C3"].value[0:10]
    result["日期"] = format_date(sheet["C3"].value[0:10])
    result["开始时间"] = sheet["C3"].value[11:16]
    result["结束时间"] = sheet["C3"].value[17:22]
    result["地点"] = sheet["J3"].value
    result["主持人"] = sheet["C4"].value
    result["记录人"] = sheet["J4"].value
    result["参会人详细信息"] = []
    result["各组出勤情况"] = {}

    rows = sheet.max_row
    get_row = "B7:" + "G" + str(rows - 1)
    # 获取参会人员信息
    cells = sheet[get_row]
    for i in cells:
        people_info = []
        total_amount += 1
        for j in i:
            people_info.append(j.value)
        result["参会人详细信息"].append(handle_info(people_info))
        if people_info[3] == "出席":
            actual_amount += 1
        else:
            absence_amount += 1
    get_row = "I7:" + "N" + str(rows - 1)
    cells = sheet[get_row]
    for i in cells:
        if i[0].value is None:
            break
        else:
            total_amount += 1
            people_info = []
            for j in i:
                people_info.append(j.value)
            result["参会人详细信息"].append(handle_info(people_info))
            if people_info[3] == "出席":
                actual_amount += 1
            else:
                absence_amount += 1

    result["应出席人数"] = total_amount
    result["实际出席人数"] = actual_amount
    result["缺席人数"] = absence_amount
    ratio = result["实际出席人数"] / result["应出席人数"]
    result["出席率"] = "%.1f%%" % (ratio * 100)

    # 各部门人员参会情况
    tongJi = {}
    for people in result["参会人详细信息"]:
        if people["部门"] in result["各组出勤情况"].keys():
            result["各组出勤情况"][people["部门"]]["应出勤人数"] += 1
        else:
            result["各组出勤情况"][people["部门"]] = {
                "应出勤人数": 1,
                "实际出勤人数": 0,
                "出勤率": 0,
                "迟到比例": 0,
                "早退比例": 0
            }
            tongJi[people["部门"]] = {
                'late_amount': 0,
                'early_amount': 0
            }

        if people["出席状态"] == "出席":
            result["各组出勤情况"][people["部门"]]["实际出勤人数"] += 1
        if people["是否迟到"]:
            tongJi[people["部门"]]["late_amount"] += 1
        if people["是否早退"]:
            tongJi[people["部门"]]["early_amount"] += 1

        ratio = result["各组出勤情况"][people["部门"]]["实际出勤人数"] / result["各组出勤情况"][people["部门"]][
            "应出勤人数"]
        if ratio != 0.0:
            result["各组出勤情况"][people["部门"]]["出勤率"] = "%.0f%%" % (ratio * 100)
        else:
            result["各组出勤情况"][people["部门"]]["出勤率"] = "0%"
            #
        if result["各组出勤情况"][people["部门"]]["实际出勤人数"] != 0:
            ratio = tongJi[people["部门"]]["late_amount"] / result["各组出勤情况"][people["部门"]]["实际出勤人数"]
        else:
            ratio = 0.0
        if ratio != 0.0:
            result["各组出勤情况"][people["部门"]]["迟到比例"] = "%.0f%%" % (ratio * 100)
        else:
            result["各组出勤情况"][people["部门"]]["迟到比例"] = "0%"
        #
        if result["各组出勤情况"][people["部门"]]["实际出勤人数"] != 0:
            ratio = tongJi[people["部门"]]["early_amount"] / result["各组出勤情况"][people["部门"]]["实际出勤人数"]
        else:
            ratio = 0.0
        if ratio != 0.0:
            result["各组出勤情况"][people["部门"]]["早退比例"] = "%.0f%%" % (ratio * 100)
        else:
            result["各组出勤情况"][people["部门"]]["早退比例"] = "0%"

    # 输出文件
    with open("bz_data.json", "w", encoding='utf-8', ) as f:
        json.dump(result, f, indent=4, sort_keys=True, ensure_ascii=False)
    write_file(result, 'bz_summary.txt')


if __name__ == '__main__':
    run("testcase/testcase1.xlsx")
    print("yes")
